#!/usr/bin/env python

import os
import re
import sys
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter  
from openpyxl.utils import get_column_letter

input_file_1 = sys.argv[1]
#/online/home/chenff/X170001-p212-testing/data/size/py/all_pos_gene_before_after
input_file_2 = sys.argv[2]
#/online/home/chenff/X170001-p212-testing/data/size/py/snp_clinvar_coms
output_file = '/online/home/chenff/X170001-p212-testing/data/size/py/all_pos_rs_before'

dict= {};n=0;m=0
with open (input_file_1,'r') as f_1:
    for pos in f_1:
        if re.match('Chr',pos):
            pos = re.split('\t',pos.strip())
            #print (len(pos));sys.exit()
            project_dict = { project_num+1 : project_tag for project_num,project_tag in enumerate(pos) }
            #print (project_dict)
            project_dict_t_n = { project_tag : project_num+1 for project_num,project_tag in enumerate(pos) }
        else:
            pos = re.split('\s',pos.strip())
            #print (pos);sys.exit() 
            format_dict = { format_tag : format_num+1 for format_num,format_tag in enumerate(pos) }
            project_name = project_dict[format_dict['+']]
            #print (format_dict);print (project_name);sys.exit()
            pos_info = ' '.join(pos[:3]);ref_alt = pos[3:6]
            if pos_info in dict.keys():
                n =n+1
                #print (pos_info)
                if project_name in dict[pos_info].keys():
                    #print (pos_info);print (project_name);sys.exit()
                    dict[pos_info][project_name].append(ref_alt)
                else:
                    dict[pos_info][project_name] = [ref_alt]
            else:
                dict[pos_info] = {project_name:[ref_alt]}
            #print (pos_info);print (ref_alt);print (dict);sys.exit()
#print (dict);sys.exit()
#print (dict['1 11722913 11722913'])
#print (dict['2 151545951'])
#print (dict['2 151545951 151545951'])

with open (input_file_2,'r') as f_2:
    for pos_rs in f_2:
        if re.search('Chr',pos_rs) ==None:
            # print (pos_rs);sys.exit()
            pos_rs_info = ' '.join(re.split('\s',pos_rs.strip())[:3])
            rs_ref_alt = [''.join(re.split('\s',pos_rs.strip())[3])]+[''.join(re.split('\s',pos_rs.strip())[4])]
            pos_rs_snp = re.split('\s',pos_rs.strip())[5];pos_rs_clinvar = re.split('\s',pos_rs.strip())[6];pos_rs_coms =re.split('\s',pos_rs.strip())[6]
            #print (pos_rs_info);print(rs_ref_alt);print (pos_rs_snp);sys.exit()
            if pos_rs_info in dict.keys():
                for name,val in dict[pos_rs_info].items():
                    #print (name);print (val);sys.exit()
                    #print (dict[pos_rs_info]);print (info);sys.exit()
                    for info in val :
                        #print (info);print(info[:2]);sys.exit()
                        if info[:2] == rs_ref_alt:
                            info.append(pos_rs_snp);info.append(pos_rs_clinvar);info.append(pos_rs_coms)
                        #print (info);sys.exit()
                        else:
                            #print (type(rs_ref_alt[1]));print (type(info[1]));sys.exit()
                            if re.search(rs_ref_alt[1],info[:2][1]):
                              # print (pos_rs_snp)
                                info.append(pos_rs_snp);info.append(pos_rs_clinvar);info.append(pos_rs_coms)
                               #print (info);print (rs_ref_alt);print(dict[pos_rs_info]);sys.exit()
                #print (dict[pos_rs_info]);sys.exit()
#print (dict['2 151545951'])
#print (dict['2 151545951 151545951']);sys.exit()
#print (dict['1 11722913 11722913']);sys.exit()
#print (dict);print (len(dict));sys.exit()

wb = Workbook()
#ew = ExcelWriter(workbook = wb)
ws = wb.worksheets[0] 
ws.title = "sheet1"
row = 1;
ws.cell(1,1).value = 'Chr';ws.cell(1,2).value = 'Start';ws.cell(1,3).value = 'End';ws.cell(1,4).value = 'Gene'
for key in dict.keys():
    key_re = re.split('\s',key)
    #print (key);print (key_re);sys.exit()
    row = row +1
    ws.cell(row,1).value = key_re[0];ws.cell(row,2).value = key_re[1];ws.cell(row,3).value = key_re[2]
    #sheet[row,0] = key_re[0];sheet[row,1] =  key_re[-1]
    #wb.save('new.xlsx');sys.exit()
    for key_k in dict[key].keys():
        name_col = project_dict_t_n[key_k]
            #print (dict[key][key_k]);sys.exit()
            #print (key);print (key_k);print(name_col);sys.exit()
            #worksheet1.write(0,name_col,key_k)
        ws.cell(1,name_col-2).value = key_k
        rs_alt_ref ='';alt_ref = ''
        for info in dict[key][key_k]:
            #print (type(info[3]));sys.exit()
            if info[3]=='.':
                alt_ref =''.join(info[0])+'>'+''.join(info[1])
                rs_alt_ref = rs_alt_ref+alt_ref+';'
                #ws.cell(row,name_col-2).value = rs_alt_ref
            else:
                alt_ref =''.join(info[0])+'>'+''.join(info[1])
                rs_alt_ref = rs_alt_ref+''.join(info[3])+': '+alt_ref+';'
            ws.cell(row,name_col-2).value = rs_alt_ref
            for str in info[2:3]:
                ws.cell(row,4).value = str
                    #worksheet1.write(row,info_col,str)
            #worksheet1.write(row,name_col,alt_ref)
#print (name_col);print (info_col);sys.exit()
wb.save('1207.xlsx')
